"""
Actualiza el Excel del MER (Modelo Entidad-Relación) para que coincida con
schema.sql del proyecto Imexcol.

Reglas:
- Solo modifica filas R7+ (datos de columnas) de las hojas de entidades.
- Preserva R1-R6 (links, cabeceras) y la hoja MER, Entidades, Valores.
"""

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

RUTA_ENTRADA = 'docs_dominio/MER_Imexcol.xlsx'
RUTA_SALIDA = 'docs_dominio/MER_Imexcol_v2.xlsx'

# Columnas del MER (1-indexed):
# 1=Nombre, 2=Tipo, 3=Longitud, 4=Enteros, 5=Decimales,
# 6=Autoincremento, 7=Obligatorio, 8=PK, 9=¿FK?, 10=Entidad FK, 11=Columna FK,
# 12=¿Unique?, 13=Nombre unique, 14=¿Check?, 15=Nombre check,
# 16=¿Indice?, 17=Tipo indice, 18=Nombre indice, 19=Alias, 20=Descripción

# Helper para construir una fila ordenada por nombre
def col(nombre, tipo, longitud=None, enteros=None, decimales=None,
        obligatorio='SI', pk='NO', fk_entidad=None, fk_columna=None,
        unique=None, indice_tipo=None, indice_nombre=None,
        descripcion=''):
    es_fk = 'SI' if fk_entidad else 'NO'
    es_unique = 'SI' if unique else 'NO'
    es_indice = 'SI' if indice_tipo else 'NO'
    return [
        nombre, tipo, longitud, enteros, decimales,
        'NO',                       # 6 autoincremento
        obligatorio,                # 7 obligatorio
        pk,                         # 8 ¿PK?
        es_fk,                      # 9 ¿FK?
        fk_entidad,                 # 10 Entidad FK
        fk_columna,                 # 11 Columna FK
        es_unique,                  # 12 ¿Unique?
        unique,                     # 13 Nombre unique
        'NO',                       # 14 ¿Check?
        None,                       # 15 Nombre check
        es_indice,                  # 16 ¿Indice?
        indice_tipo,                # 17 Tipo indice
        indice_nombre,              # 18 Nombre indice
        nombre,                     # 19 Alias = nombre por defecto
        descripcion,                # 20 Descripción
    ]


# ============================================================
# Definición de tablas según schema.sql
# ============================================================
TABLAS = {
    'Cliente': {
        'tabla_sql': 'imex_cliente',
        'columnas': [
            col('id', 'UNIQUEIDENTIFIER', pk='SI',
                descripcion='Identificador único universal del cliente.'),
            col('nombre', 'NVARCHAR', longitud=100,
                descripcion='Nombre o nombres del cliente.'),
            col('apellido', 'NVARCHAR', longitud=100,
                descripcion='Apellido o apellidos del cliente.'),
            col('correo_electronico', 'NVARCHAR', longitud=150,
                unique='uq_imex_cliente_correo',
                descripcion='Correo electrónico único de contacto y autenticación.'),
            col('contrasena', 'NVARCHAR', longitud=255,
                descripcion='Contraseña del cliente almacenada de forma segura.'),
            col('telefono', 'NVARCHAR', longitud=20, obligatorio='NO',
                descripcion='Número de teléfono de contacto (opcional).'),
            col('estado', 'BIT',
                descripcion='Indica si el cliente está activo (1) o inactivo (0).'),
        ],
    },
    'Administrador': {
        'tabla_sql': 'imex_administrador',
        'columnas': [
            col('id', 'UNIQUEIDENTIFIER', pk='SI',
                descripcion='Identificador único universal del administrador.'),
            col('nombre_usuario', 'NVARCHAR', longitud=50,
                unique='uq_imex_administrador_nombre_usuario',
                descripcion='Nombre de usuario único del administrador.'),
            col('correo_electronico', 'NVARCHAR', longitud=150,
                unique='uq_imex_administrador_correo',
                descripcion='Correo electrónico único del administrador.'),
            col('contrasena', 'NVARCHAR', longitud=255,
                descripcion='Contraseña del administrador almacenada de forma segura.'),
            col('estado', 'BIT',
                descripcion='Indica si el administrador está activo (1) o inactivo (0).'),
        ],
    },
    'Direccion': {
        'tabla_sql': 'imex_direccion',
        'columnas': [
            col('id', 'UNIQUEIDENTIFIER', pk='SI',
                descripcion='Identificador único universal de la dirección.'),
            col('cliente_id', 'UNIQUEIDENTIFIER',
                fk_entidad='Cliente', fk_columna='id',
                descripcion='Referencia al cliente al que pertenece la dirección.'),
            col('calle', 'NVARCHAR', longitud=200,
                descripcion='Nomenclatura exacta de la dirección.'),
            col('ciudad', 'NVARCHAR', longitud=100,
                descripcion='Ciudad o municipio donde se ubica la dirección.'),
            col('departamento', 'NVARCHAR', longitud=100,
                descripcion='Departamento o estado donde se ubica la dirección.'),
            col('codigo_postal', 'NVARCHAR', longitud=10, obligatorio='NO',
                descripcion='Código postal de la dirección (opcional).'),
            col('pais', 'NVARCHAR', longitud=100,
                descripcion='País donde se ubica la dirección.'),
            col('es_principal', 'BIT',
                descripcion='Indica si es la dirección principal del cliente.'),
        ],
    },
    'Categoria': {
        'tabla_sql': 'imex_categoria',
        'columnas': [
            col('id', 'UNIQUEIDENTIFIER', pk='SI',
                descripcion='Identificador único universal de la categoría.'),
            col('nombre', 'NVARCHAR', longitud=100,
                descripcion='Nombre comercial de la categoría.'),
            col('descripcion', 'NVARCHAR', longitud=255, obligatorio='NO',
                descripcion='Descripción breve de la categoría (opcional).'),
            col('estado', 'BIT',
                descripcion='Indica si la categoría está activa (1) o inactiva (0).'),
        ],
    },
    'Producto': {
        'tabla_sql': 'imex_producto',
        'columnas': [
            col('id', 'UNIQUEIDENTIFIER', pk='SI',
                descripcion='Identificador único universal del producto.'),
            col('categoria_id', 'UNIQUEIDENTIFIER',
                fk_entidad='Categoria', fk_columna='id',
                indice_tipo='ASC', indice_nombre='ix_imex_producto_categoria_id',
                descripcion='Referencia a la categoría a la que pertenece el producto.'),
            col('nombre', 'NVARCHAR', longitud=100,
                descripcion='Nombre comercial del producto.'),
            col('descripcion', 'NVARCHAR', longitud=1000, obligatorio='NO',
                descripcion='Descripción detallada del producto (opcional).'),
            col('precio', 'DECIMAL', enteros=12, decimales=2,
                descripcion='Precio del producto en pesos colombianos (COP).'),
            col('stock', 'INT',
                descripcion='Cantidad de unidades disponibles en inventario.'),
            col('imagen_url', 'NVARCHAR', longitud=255, obligatorio='NO',
                descripcion='Ruta relativa o URL de la imagen del producto (opcional).'),
            col('estado', 'BIT',
                descripcion='Indica si el producto está activo (1) o inactivo (0).'),
        ],
    },
    'CarritoCompras': {
        'tabla_sql': 'imex_carrito_compras',
        'columnas': [
            col('id', 'UNIQUEIDENTIFIER', pk='SI',
                descripcion='Identificador único universal del carrito.'),
            col('cliente_id', 'UNIQUEIDENTIFIER',
                fk_entidad='Cliente', fk_columna='id',
                descripcion='Referencia al cliente propietario del carrito.'),
            col('fecha_creacion', 'DATE',
                descripcion='Fecha de creación del carrito.'),
            col('estado', 'NVARCHAR', longitud=20,
                descripcion='Estado del carrito (ACTIVO, ABANDONADO, FINALIZADO).'),
        ],
    },
    'ItemCarrito': {
        'tabla_sql': 'imex_item_carrito',
        'columnas': [
            col('id', 'UNIQUEIDENTIFIER', pk='SI',
                descripcion='Identificador único universal del ítem del carrito.'),
            col('carrito_id', 'UNIQUEIDENTIFIER',
                fk_entidad='CarritoCompras', fk_columna='id',
                descripcion='Referencia al carrito al que pertenece el ítem.'),
            col('producto_id', 'UNIQUEIDENTIFIER',
                fk_entidad='Producto', fk_columna='id',
                descripcion='Referencia al producto agregado al carrito.'),
            col('cantidad', 'INT',
                descripcion='Número de unidades del producto en el carrito.'),
            col('precio_unitario', 'DECIMAL', enteros=12, decimales=2,
                descripcion='Precio unitario congelado del producto en el carrito.'),
        ],
    },
    'Pedido': {
        'tabla_sql': 'imex_pedido',
        'columnas': [
            col('id', 'UNIQUEIDENTIFIER', pk='SI',
                descripcion='Identificador único universal del pedido.'),
            col('cliente_id', 'UNIQUEIDENTIFIER',
                fk_entidad='Cliente', fk_columna='id',
                indice_tipo='ASC', indice_nombre='ix_imex_pedido_cliente_id',
                descripcion='Referencia al cliente que realizó el pedido.'),
            col('fecha_pedido', 'DATE',
                descripcion='Fecha en que se generó el pedido.'),
            col('total', 'DECIMAL', enteros=12, decimales=2,
                descripcion='Valor total del pedido en pesos colombianos (COP).'),
            col('estado', 'NVARCHAR', longitud=50,
                indice_tipo='ASC', indice_nombre='ix_imex_pedido_estado',
                descripcion='Estado del pedido (PENDIENTE, CONFIRMADO, PREPARANDO, ENVIADO, ENTREGADO, CANCELADO).'),
            col('direccion_id', 'UNIQUEIDENTIFIER', obligatorio='NO',
                fk_entidad='Direccion', fk_columna='id',
                descripcion='Referencia a la dirección de envío del pedido (opcional).'),
        ],
    },
    'LineaPedido': {
        'tabla_sql': 'imex_linea_pedido',
        'columnas': [
            col('id', 'UNIQUEIDENTIFIER', pk='SI',
                descripcion='Identificador único universal de la línea de pedido.'),
            col('pedido_id', 'UNIQUEIDENTIFIER',
                fk_entidad='Pedido', fk_columna='id',
                descripcion='Referencia al pedido al que pertenece la línea.'),
            col('producto_id', 'UNIQUEIDENTIFIER',
                fk_entidad='Producto', fk_columna='id',
                descripcion='Referencia al producto incluido en la línea.'),
            col('cantidad', 'INT',
                descripcion='Número de unidades del producto en la línea.'),
            col('precio_unitario', 'DECIMAL', enteros=12, decimales=2,
                descripcion='Precio unitario congelado del producto al momento del pedido.'),
            col('subtotal', 'DECIMAL', enteros=12, decimales=2,
                descripcion='Subtotal de la línea (cantidad × precioUnitario).'),
        ],
    },
    'Pago': {
        'tabla_sql': 'imex_pago',
        'columnas': [
            col('id', 'UNIQUEIDENTIFIER', pk='SI',
                descripcion='Identificador único universal del pago.'),
            col('pedido_id', 'UNIQUEIDENTIFIER',
                fk_entidad='Pedido', fk_columna='id',
                descripcion='Referencia al pedido asociado al pago.'),
            col('metodo_pago_id', 'UNIQUEIDENTIFIER',
                fk_entidad='MetodoPago', fk_columna='id',
                descripcion='Referencia al método de pago utilizado.'),
            col('monto', 'DECIMAL', enteros=12, decimales=2,
                descripcion='Monto del pago en pesos colombianos (COP).'),
            col('fecha_pago', 'DATE',
                descripcion='Fecha en que se procesó el pago.'),
            col('estado', 'NVARCHAR', longitud=50,
                descripcion='Estado del pago (PENDIENTE, EN_PROCESO, APROBADO, RECHAZADO, REEMBOLSADO).'),
        ],
    },
    'MetodoPago': {
        'tabla_sql': 'imex_metodo_pago',
        'columnas': [
            col('id', 'UNIQUEIDENTIFIER', pk='SI',
                descripcion='Identificador único universal del método de pago.'),
            col('nombre', 'NVARCHAR', longitud=100,
                descripcion='Nombre comercial del método de pago.'),
            col('descripcion', 'NVARCHAR', longitud=255, obligatorio='NO',
                descripcion='Descripción del método de pago (opcional).'),
            col('estado', 'BIT',
                descripcion='Indica si el método de pago está activo (1) o inactivo (0).'),
        ],
    },
    'Envio': {
        'tabla_sql': 'imex_envio',
        'columnas': [
            col('id', 'UNIQUEIDENTIFIER', pk='SI',
                descripcion='Identificador único universal del envío.'),
            col('pedido_id', 'UNIQUEIDENTIFIER',
                fk_entidad='Pedido', fk_columna='id',
                descripcion='Referencia al pedido al que pertenece el envío.'),
            col('fecha_envio', 'DATE',
                descripcion='Fecha en que se despachó el envío.'),
            col('fecha_entrega', 'DATE', obligatorio='NO',
                descripcion='Fecha estimada o real de entrega del envío (opcional).'),
            col('transportadora', 'NVARCHAR', longitud=100,
                descripcion='Nombre de la empresa transportadora.'),
            col('numero_guia', 'NVARCHAR', longitud=100,
                descripcion='Número de guía proporcionado por la transportadora.'),
            col('estado', 'NVARCHAR', longitud=50,
                descripcion='Estado del envío (PREPARANDO, EN_TRANSITO, ENTREGADO, DEVUELTO, EXTRAVIADO).'),
        ],
    },
}

# Hojas a NO tocar (no tienen tabla en schema.sql)
HOJAS_INTACTAS = {'ImagenProducto', 'Resena', 'Factura', 'EventoRastreo'}


def desunir_si_combinada(ws, fila, col_idx):
    coord = ws.cell(row=fila, column=col_idx).coordinate
    for rango in list(ws.merged_cells.ranges):
        min_col, min_row, max_col, max_row = range_boundaries(str(rango))
        if min_row <= fila <= max_row and min_col <= col_idx <= max_col:
            ws.unmerge_cells(str(rango))
            return True
    return False


def escribir(ws, fila, col_idx, valor):
    desunir_si_combinada(ws, fila, col_idx)
    ws.cell(row=fila, column=col_idx).value = valor


def actualizar_hoja_mer(ws, definicion):
    fila_inicio = 7  # primera fila de datos en hojas MER
    columnas = definicion['columnas']
    num_filas_nuevas = len(columnas)

    # Detectar última fila con datos previos (para limpiar las que sobren)
    fila_fin_anterior = ws.max_row

    # Escribir las columnas
    for i, fila_datos in enumerate(columnas):
        fila_excel = fila_inicio + i
        for c, valor in enumerate(fila_datos, start=1):
            escribir(ws, fila_excel, c, valor)

    # Limpiar filas extra que tuviera la hoja previa
    fila_fin_nueva = fila_inicio + num_filas_nuevas - 1
    if fila_fin_anterior > fila_fin_nueva:
        for fila in range(fila_fin_nueva + 1, fila_fin_anterior + 1):
            for c in range(1, 21):
                desunir_si_combinada(ws, fila, c)
                ws.cell(row=fila, column=c).value = None

    print(f'  [OK] {ws.title}: {num_filas_nuevas} columnas SQL escritas '
          f'({fila_inicio}-{fila_fin_nueva})')


def main():
    print(f'Cargando {RUTA_ENTRADA}…')
    wb = load_workbook(RUTA_ENTRADA)
    print(f'  Hojas: {len(wb.sheetnames)}')

    print('\n=== Actualizando hojas de entidades ===')
    for hoja_nombre, definicion in TABLAS.items():
        if hoja_nombre not in wb.sheetnames:
            print(f'  [NO EXISTE] {hoja_nombre}')
            continue
        actualizar_hoja_mer(wb[hoja_nombre], definicion)

    print('\n=== Hojas intactas (sin tabla SQL) ===')
    for nombre in HOJAS_INTACTAS:
        if nombre in wb.sheetnames:
            print(f'  [SKIP] {nombre}')

    print(f'\nGuardando {RUTA_SALIDA}…')
    wb.save(RUTA_SALIDA)
    print('Listo.')


if __name__ == '__main__':
    main()
