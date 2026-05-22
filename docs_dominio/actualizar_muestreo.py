"""
Actualiza el Excel de Muestreo de Datos del Modelo de Dominio para que coincida
con el código Java actual del proyecto Imexcol.

Reglas:
- Solo modifica filas de descripciones, nombres y 5 filas de datos en hojas M-.
- NO toca hojas C-, E-, ni del modelo de dominio, ni ListadoObjetosDominio.
- NO toca leyenda de colores (filas posteriores a los datos).
- Preserva formato (colores, bordes, fuentes) de las celdas existentes.
"""

from copy import copy
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

RUTA_ENTRADA = 'docs_dominio/MuestreoDatosModeloDominio Imexcol.xlsx'
RUTA_SALIDA = 'docs_dominio/MuestreoDatosModeloDominio_Imexcol_v2.xlsx'

# ============================================================
# Definición de atributos por hoja, según el código Java actual.
# Para cada hoja: tipos del atributo + descripción + 5 datos de ejemplo.
# ============================================================

# UUIDs realistas estables — formato estilo Microsoft (mayúsculas)
UUID_CLI = [
    '3F2504E0-4F89-11D3-9A0C-0305E82C3301',
    '3F2504E0-4F89-11D3-9A0C-0305E82C3302',
    '3F2504E0-4F89-11D3-9A0C-0305E82C3303',
    '3F2504E0-4F89-11D3-9A0C-0305E82C3304',
    '3F2504E0-4F89-11D3-9A0C-0305E82C3305',
]
UUID_ADM = [
    'A1B2C3D4-5566-4F77-8899-AABBCCDDEE01',
    'A1B2C3D4-5566-4F77-8899-AABBCCDDEE02',
    'A1B2C3D4-5566-4F77-8899-AABBCCDDEE03',
    'A1B2C3D4-5566-4F77-8899-AABBCCDDEE04',
    'A1B2C3D4-5566-4F77-8899-AABBCCDDEE05',
]
UUID_DIR = [
    'D1234567-8901-4234-5678-90123456D001',
    'D1234567-8901-4234-5678-90123456D002',
    'D1234567-8901-4234-5678-90123456D003',
    'D1234567-8901-4234-5678-90123456D004',
    'D1234567-8901-4234-5678-90123456D005',
]
UUID_CAT = [
    'CA111111-1111-4111-8111-111111111101',
    'CA111111-1111-4111-8111-111111111102',
    'CA111111-1111-4111-8111-111111111103',
    'CA111111-1111-4111-8111-111111111104',
    'CA111111-1111-4111-8111-111111111105',
]
UUID_PRO = [
    '4F8A0001-2222-4333-8444-555566667701',
    '4F8A0001-2222-4333-8444-555566667702',
    '4F8A0001-2222-4333-8444-555566667703',
    '4F8A0001-2222-4333-8444-555566667704',
    '4F8A0001-2222-4333-8444-555566667705',
]
UUID_CAR = [
    'CC000001-AAAA-4BBB-8CCC-DDDDEEEEFF01',
    'CC000001-AAAA-4BBB-8CCC-DDDDEEEEFF02',
    'CC000001-AAAA-4BBB-8CCC-DDDDEEEEFF03',
    'CC000001-AAAA-4BBB-8CCC-DDDDEEEEFF04',
    'CC000001-AAAA-4BBB-8CCC-DDDDEEEEFF05',
]
UUID_ITM = [
    '1C100000-AAAA-4BBB-8CCC-000000000001',
    '1C100000-AAAA-4BBB-8CCC-000000000002',
    '1C100000-AAAA-4BBB-8CCC-000000000003',
    '1C100000-AAAA-4BBB-8CCC-000000000004',
    '1C100000-AAAA-4BBB-8CCC-000000000005',
]
UUID_PED = [
    'PED12345-A1B2-4C3D-8E4F-000000000001',
    'PED12345-A1B2-4C3D-8E4F-000000000002',
    'PED12345-A1B2-4C3D-8E4F-000000000003',
    'PED12345-A1B2-4C3D-8E4F-000000000004',
    'PED12345-A1B2-4C3D-8E4F-000000000005',
]
UUID_LIN = [
    'LP000000-1111-4222-8333-000000000001',
    'LP000000-1111-4222-8333-000000000002',
    'LP000000-1111-4222-8333-000000000003',
    'LP000000-1111-4222-8333-000000000004',
    'LP000000-1111-4222-8333-000000000005',
]
UUID_MET = [
    'MP000000-1111-4222-8333-AAAAAA000001',
    'MP000000-1111-4222-8333-AAAAAA000002',
    'MP000000-1111-4222-8333-AAAAAA000003',
    'MP000000-1111-4222-8333-AAAAAA000004',
    'MP000000-1111-4222-8333-AAAAAA000005',
]
UUID_PAG = [
    'PA000000-1111-4222-8333-BBBBBB000001',
    'PA000000-1111-4222-8333-BBBBBB000002',
    'PA000000-1111-4222-8333-BBBBBB000003',
    'PA000000-1111-4222-8333-BBBBBB000004',
    'PA000000-1111-4222-8333-BBBBBB000005',
]
UUID_ENV = [
    'EN000000-1111-4222-8333-CCCCCC000001',
    'EN000000-1111-4222-8333-CCCCCC000002',
    'EN000000-1111-4222-8333-CCCCCC000003',
    'EN000000-1111-4222-8333-CCCCCC000004',
    'EN000000-1111-4222-8333-CCCCCC000005',
]


def vp(s):
    return s  # placeholder semántico


HOJAS = {
    # --- M-Cliente ---
    'M-Cliente': {
        'nombres': ['id', 'nombre', 'apellido', 'correoElectronico',
                    'contrasena', 'telefono', 'estado'],
        'descripciones': [
            'Identificador único universal (UUID) del cliente, generado automáticamente al registrarse.',
            'Representa el nombre o nombres del cliente.',
            'Representa el apellido o apellidos del cliente.',
            'Representa el correo electrónico único de contacto y autenticación del cliente.',
            'Contraseña del cliente almacenada de forma segura para autenticación.',
            'Número de teléfono de contacto del cliente (opcional).',
            'Indica si el cliente está activo (true) o inactivo (false) en el sistema.',
        ],
        'datos': [
            [UUID_CLI[0], 'Carlos', 'Restrepo', 'carlos.restrepo@imexcol.com',
             '********', '3001112233', True],
            [UUID_CLI[1], 'María', 'González', 'maria.gonzalez@imexcol.com',
             '********', '3104445566', True],
            [UUID_CLI[2], 'Andrés', 'Mejía', 'andres.mejia@imexcol.com',
             '********', '3207778899', True],
            [UUID_CLI[3], 'Camila', 'Ramírez', 'camila.ramirez@imexcol.com',
             '********', '3151231234', True],
            [UUID_CLI[4], 'Daniel', 'Pérez', 'daniel.perez@imexcol.com',
             '********', '3009998877', False],
        ],
    },

    # --- M-Administrador ---
    'M-Administrador': {
        'nombres': ['id', 'nombreUsuario', 'correoElectronico',
                    'contrasena', 'estado'],
        'descripciones': [
            'Identificador único universal (UUID) del administrador.',
            'Nombre de usuario único del administrador para iniciar sesión.',
            'Correo electrónico de contacto del administrador.',
            'Contraseña del administrador almacenada de forma segura.',
            'Indica si el administrador está activo (true) o inactivo (false).',
        ],
        'datos': [
            [UUID_ADM[0], 'admin.principal', 'admin@imexcol.com', '********', True],
            [UUID_ADM[1], 'lmartinez', 'l.martinez@imexcol.com', '********', True],
            [UUID_ADM[2], 'sgomez', 's.gomez@imexcol.com', '********', True],
            [UUID_ADM[3], 'alopez', 'a.lopez@imexcol.com', '********', False],
            [UUID_ADM[4], 'mcastro', 'm.castro@imexcol.com', '********', True],
        ],
    },

    # --- M-Dirección ---
    'M-Dirección': {
        'nombres': ['id', 'cliente', 'calle', 'ciudad', 'departamento',
                    'pais', 'codigoPostal', 'esPrincipal'],
        'descripciones': [
            'Identificador único universal (UUID) de la dirección.',
            'Referencia al cliente al que pertenece esta dirección.',
            'Nomenclatura exacta de la dirección (calle, carrera, número).',
            'Ciudad o municipio donde se ubica la dirección.',
            'Departamento o estado donde se ubica la dirección.',
            'País donde se ubica la dirección.',
            'Código postal de la dirección (opcional).',
            'Indica si esta es la dirección principal del cliente (true/false).',
        ],
        'datos': [
            [UUID_DIR[0], UUID_CLI[0], 'Calle 10 # 43-20',
             'Medellín', 'Antioquia', 'Colombia', '050021', True],
            [UUID_DIR[1], UUID_CLI[1], 'Carrera 48 # 32-15',
             'Envigado', 'Antioquia', 'Colombia', '055422', False],
            [UUID_DIR[2], UUID_CLI[2], 'Carrera 15 # 85-30',
             'Bogotá', 'Cundinamarca', 'Colombia', '110221', True],
            [UUID_DIR[3], UUID_CLI[3], 'Avenida 6N # 24-10',
             'Cali', 'Valle del Cauca', 'Colombia', '760045', True],
            [UUID_DIR[4], UUID_CLI[4], 'Calle 84 # 51B-20',
             'Barranquilla', 'Atlántico', 'Colombia', '080020', False],
        ],
    },

    # --- M-Categoria ---
    'M-Categoria': {
        'nombres': ['id', 'nombre', 'descripcion', 'estado'],
        'descripciones': [
            'Identificador único universal (UUID) de la categoría.',
            'Nombre comercial de la categoría de productos.',
            'Descripción breve de la categoría (opcional).',
            'Indica si la categoría está activa (true) o inactiva (false).',
        ],
        'datos': [
            [UUID_CAT[0], 'Hogar', 'Productos decorativos y utilitarios para el hogar.', True],
            [UUID_CAT[1], 'Moda', 'Prendas de vestir y accesorios de moda.', True],
            [UUID_CAT[2], 'Accesorios', 'Accesorios personales y complementos.', True],
            [UUID_CAT[3], 'Tecnología', 'Dispositivos y accesorios tecnológicos.', True],
            [UUID_CAT[4], 'Cocina', 'Utensilios y menaje de cocina.', False],
        ],
    },

    # --- M-Producto ---
    'M-Producto': {
        'nombres': ['id', 'categoria', 'nombre', 'descripcion',
                    'precio', 'stock', 'imagenUrl', 'estado'],
        'descripciones': [
            'Identificador único universal (UUID) del producto.',
            'Referencia a la categoría a la que pertenece el producto.',
            'Nombre comercial del producto.',
            'Descripción detallada de las características del producto.',
            'Precio del producto en pesos colombianos (COP).',
            'Cantidad de unidades disponibles en inventario.',
            'Ruta relativa a la imagen representativa del producto (opcional).',
            'Indica si el producto está activo (true) o inactivo (false).',
        ],
        'datos': [
            [UUID_PRO[0], UUID_CAT[0], 'Lámpara decorativa NP5115',
             'Lámpara de mesa decorativa con base metálica y pantalla en tela.',
             89900, 25, 'img/Lampara_NP5115.png', True],
            [UUID_PRO[1], UUID_CAT[2], 'Sombrilla antiviento NP4292',
             'Sombrilla plegable antiviento con forro reforzado y mango ergonómico.',
             145000, 40, 'img/Sombrilla_antiviento_NP4292.png', True],
            [UUID_PRO[2], UUID_CAT[1], 'Gorra para hombre NP5275',
             'Gorra deportiva para hombre con visera curva y ajuste trasero.',
             35500, 60, 'img/Gorra_hombre_NP5275.png', True],
            [UUID_PRO[3], UUID_CAT[4], 'Set de cucharas inoxidables x 12',
             'Set de 12 cucharas en acero inoxidable con diseño labrado.',
             32000, 0, '', False],
            [UUID_PRO[4], UUID_CAT[3], 'Audífonos inalámbricos',
             'Audífonos bluetooth con cancelación de ruido y estuche de carga.',
             189900, 15, '', True],
        ],
    },

    # --- M-CarritoCompras ---
    'M-CarritoCompras': {
        'nombres': ['id', 'cliente', 'fechaCreacion', 'estado'],
        'descripciones': [
            'Identificador único universal (UUID) del carrito de compras.',
            'Referencia al cliente propietario del carrito.',
            'Fecha de creación del carrito (LocalDate).',
            'Estado del carrito (ACTIVO, ABANDONADO, FINALIZADO).',
        ],
        'datos': [
            [UUID_CAR[0], UUID_CLI[0], '2026-05-18', 'ACTIVO'],
            [UUID_CAR[1], UUID_CLI[1], '2026-05-19', 'FINALIZADO'],
            [UUID_CAR[2], UUID_CLI[2], '2026-05-19', 'ACTIVO'],
            [UUID_CAR[3], UUID_CLI[3], '2026-05-20', 'ABANDONADO'],
            [UUID_CAR[4], UUID_CLI[4], '2026-05-21', 'ACTIVO'],
        ],
    },

    # --- M-ItemCarrito ---
    'M-ItemCarrito': {
        'nombres': ['id', 'carritoCompras', 'producto', 'cantidad',
                    'precioUnitario', 'subtotal'],
        'descripciones': [
            'Identificador único universal (UUID) del ítem del carrito.',
            'Referencia al carrito de compras al que pertenece el ítem.',
            'Referencia al producto agregado al carrito.',
            'Número de unidades del producto en el carrito.',
            'Precio unitario congelado del producto al momento de agregarlo.',
            'Subtotal del ítem (cantidad × precioUnitario).',
        ],
        'datos': [
            [UUID_ITM[0], UUID_CAR[0], UUID_PRO[0], 1, 89900, 89900],
            [UUID_ITM[1], UUID_CAR[1], UUID_PRO[1], 2, 145000, 290000],
            [UUID_ITM[2], UUID_CAR[2], UUID_PRO[2], 3, 35500, 106500],
            [UUID_ITM[3], UUID_CAR[3], UUID_PRO[4], 1, 189900, 189900],
            [UUID_ITM[4], UUID_CAR[4], UUID_PRO[0], 2, 89900, 179800],
        ],
    },

    # --- M-Pedido ---
    'M-Pedido': {
        'nombres': ['id', 'cliente', 'fechaPedido', 'total', 'estado'],
        'descripciones': [
            'Identificador único universal (UUID) del pedido.',
            'Referencia al cliente que realizó el pedido.',
            'Fecha en que se generó el pedido (LocalDate).',
            'Valor total del pedido en pesos colombianos (COP).',
            'Estado del pedido (PENDIENTE, CONFIRMADO, PREPARANDO, ENVIADO, ENTREGADO, CANCELADO).',
        ],
        'datos': [
            [UUID_PED[0], UUID_CLI[0], '2026-05-15', 89900, 'PENDIENTE'],
            [UUID_PED[1], UUID_CLI[1], '2026-05-16', 290000, 'CONFIRMADO'],
            [UUID_PED[2], UUID_CLI[2], '2026-05-17', 106500, 'PREPARANDO'],
            [UUID_PED[3], UUID_CLI[3], '2026-05-18', 189900, 'ENVIADO'],
            [UUID_PED[4], UUID_CLI[4], '2026-05-19', 179800, 'ENTREGADO'],
        ],
    },

    # --- M-LineaPedido ---
    'M-LineaPedido': {
        'nombres': ['id', 'pedido', 'producto', 'cantidad',
                    'precioUnitario', 'subtotal'],
        'descripciones': [
            'Identificador único universal (UUID) de la línea de pedido.',
            'Referencia al pedido al que pertenece esta línea.',
            'Referencia al producto incluido en esta línea.',
            'Número de unidades del producto en la línea.',
            'Precio unitario congelado del producto al momento del pedido.',
            'Subtotal de la línea (cantidad × precioUnitario).',
        ],
        'datos': [
            [UUID_LIN[0], UUID_PED[0], UUID_PRO[0], 1, 89900, 89900],
            [UUID_LIN[1], UUID_PED[1], UUID_PRO[1], 2, 145000, 290000],
            [UUID_LIN[2], UUID_PED[2], UUID_PRO[2], 3, 35500, 106500],
            [UUID_LIN[3], UUID_PED[3], UUID_PRO[4], 1, 189900, 189900],
            [UUID_LIN[4], UUID_PED[4], UUID_PRO[0], 2, 89900, 179800],
        ],
    },

    # --- M-Pagos ---
    'M-Pagos': {
        'nombres': ['id', 'pedido', 'metodoPago', 'monto',
                    'fechaPago', 'estado'],
        'descripciones': [
            'Identificador único universal (UUID) del pago.',
            'Referencia al pedido asociado al pago.',
            'Referencia al método de pago utilizado.',
            'Monto del pago en pesos colombianos (COP).',
            'Fecha en que se procesó el pago (LocalDate).',
            'Estado del pago (PENDIENTE, EN_PROCESO, APROBADO, RECHAZADO, REEMBOLSADO).',
        ],
        'datos': [
            [UUID_PAG[0], UUID_PED[0], UUID_MET[0], 89900,  '2026-05-15', 'PENDIENTE'],
            [UUID_PAG[1], UUID_PED[1], UUID_MET[1], 290000, '2026-05-16', 'APROBADO'],
            [UUID_PAG[2], UUID_PED[2], UUID_MET[2], 106500, '2026-05-17', 'EN_PROCESO'],
            [UUID_PAG[3], UUID_PED[3], UUID_MET[1], 189900, '2026-05-18', 'APROBADO'],
            [UUID_PAG[4], UUID_PED[4], UUID_MET[3], 179800, '2026-05-19', 'RECHAZADO'],
        ],
    },

    # --- M-MetodoPago ---
    'M-MetodoPago': {
        'nombres': ['id', 'nombre', 'descripcion', 'estado'],
        'descripciones': [
            'Identificador único universal (UUID) del método de pago.',
            'Nombre comercial del método de pago.',
            'Descripción del método de pago (opcional).',
            'Indica si el método de pago está activo (true) o inactivo (false).',
        ],
        'datos': [
            [UUID_MET[0], 'Tarjeta de crédito',   'Visa, Mastercard, American Express.', True],
            [UUID_MET[1], 'PSE',                  'Pago seguro en línea débito bancos colombianos.', True],
            [UUID_MET[2], 'Nequi',                'Billetera digital Bancolombia.', True],
            [UUID_MET[3], 'Efectivo contraentrega','Pago en efectivo al recibir el pedido.', True],
            [UUID_MET[4], 'Daviplata',            'Billetera digital Davivienda.', False],
        ],
    },

    # --- M-Envio ---
    'M-Envio': {
        'nombres': ['id', 'pedido', 'transportadora', 'numeroGuia',
                    'fechaEnvio', 'fechaEntrega', 'estado'],
        'descripciones': [
            'Identificador único universal (UUID) del envío.',
            'Referencia al pedido al que pertenece el envío.',
            'Nombre de la empresa transportadora encargada del envío.',
            'Número de guía proporcionado por la transportadora.',
            'Fecha en que se despachó el envío (LocalDate).',
            'Fecha estimada o real de entrega del envío (LocalDate, opcional).',
            'Estado del envío (PREPARANDO, EN_TRANSITO, ENTREGADO, DEVUELTO, EXTRAVIADO).',
        ],
        'datos': [
            [UUID_ENV[0], UUID_PED[1], 'Servientrega',     'SERV-30012345678', '2026-05-17', '2026-05-19', 'EN_TRANSITO'],
            [UUID_ENV[1], UUID_PED[2], 'Coordinadora',     'COOR-21456789012', '2026-05-18', '',           'PREPARANDO'],
            [UUID_ENV[2], UUID_PED[3], 'Inter Rapídisimo', 'INTE-70009876543', '2026-05-19', '2026-05-20', 'ENTREGADO'],
            [UUID_ENV[3], UUID_PED[4], 'TCC',              'TCC-54321987654',  '2026-05-19', '',           'EN_TRANSITO'],
            [UUID_ENV[4], UUID_PED[0], 'Envia',            'ENVI-09876543210', '2026-05-20', '',           'PREPARANDO'],
        ],
    },
}


def encontrar_layout(ws):
    """Detecta dinámicamente las filas de descripciones, nombres y datos."""
    for r in range(1, 12):
        v = ws.cell(row=r, column=1).value
        if isinstance(v, str) and v.strip().rstrip(':').lower() == 'descripción':
            return {
                'descripciones': r + 1,
                'nombres':       r + 2,
                'datos_inicio':  r + 3,
                'datos_fin':     r + 7,  # 5 filas inclusive
            }
    return None


def desunir_si_combinada(ws, fila, col):
    """Si la celda forma parte de un rango combinado, lo desune para poder escribir."""
    from openpyxl.utils.cell import range_boundaries
    objetivo = ws.cell(row=fila, column=col).coordinate
    for rango in list(ws.merged_cells.ranges):
        min_col, min_row, max_col, max_row = range_boundaries(str(rango))
        if min_row <= fila <= max_row and min_col <= col <= max_col:
            ws.unmerge_cells(str(rango))
            return True
    return False


def escribir(ws, fila, col, valor):
    desunir_si_combinada(ws, fila, col)
    ws.cell(row=fila, column=col).value = valor


def limpiar_celda_valor(ws, fila, col):
    """Borra solo el valor de la celda, manteniendo su formato."""
    desunir_si_combinada(ws, fila, col)
    ws.cell(row=fila, column=col).value = None


def actualizar_hoja(ws, definicion):
    layout = encontrar_layout(ws)
    if layout is None:
        print(f'  [SKIP] {ws.title}: no se encontró fila "Descripción"')
        return

    nombres = definicion['nombres']
    descripciones = definicion['descripciones']
    datos = definicion['datos']

    num_cols_nuevas = len(nombres)
    max_cols_existentes = ws.max_column

    # Sobrescribir descripciones y nombres en las columnas nuevas
    for idx, (desc, nom) in enumerate(zip(descripciones, nombres), start=1):
        escribir(ws, layout['descripciones'], idx, desc)
        escribir(ws, layout['nombres'], idx, nom)

    # Limpiar columnas adicionales en la fila de descripciones/nombres
    # (la hoja original puede tener "Combinación única" en columnas extra)
    for col in range(num_cols_nuevas + 1, max_cols_existentes + 1):
        limpiar_celda_valor(ws, layout['descripciones'], col)
        limpiar_celda_valor(ws, layout['nombres'], col)

    # Sobrescribir las 5 filas de datos
    for i, fila_datos in enumerate(datos):
        fila_excel = layout['datos_inicio'] + i
        for idx, valor in enumerate(fila_datos, start=1):
            escribir(ws, fila_excel, idx, valor)
        # Limpiar columnas extra de esa fila
        for col in range(len(fila_datos) + 1, max_cols_existentes + 1):
            limpiar_celda_valor(ws, fila_excel, col)

    print(f'  [OK] {ws.title}: {num_cols_nuevas} cols × 5 filas actualizadas '
          f'(desc fila {layout["descripciones"]}, '
          f'nombres fila {layout["nombres"]}, '
          f'datos {layout["datos_inicio"]}-{layout["datos_fin"]})')


def main():
    print(f'Cargando {RUTA_ENTRADA}…')
    wb = load_workbook(RUTA_ENTRADA)
    print(f'  Hojas: {len(wb.sheetnames)}')

    print('\n=== Actualizando hojas M- ===')
    for hoja_nombre, definicion in HOJAS.items():
        if hoja_nombre not in wb.sheetnames:
            print(f'  [NO EXISTE] {hoja_nombre}')
            continue
        actualizar_hoja(wb[hoja_nombre], definicion)

    print(f'\nGuardando {RUTA_SALIDA}…')
    wb.save(RUTA_SALIDA)
    print('Listo.')


if __name__ == '__main__':
    main()
