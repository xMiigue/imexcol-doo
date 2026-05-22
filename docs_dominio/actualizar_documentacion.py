"""
Actualiza la hoja "Clases" del Excel de Documentación de Clases para que liste
los 252 archivos .java actuales del proyecto Imexcol.

NO toca las hojas de dominio individuales (Cliente, Producto, etc.) ni el
Diagrama de clases ni Valores.
"""

import os
import re
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

RUTA_ENTRADA = 'docs_dominio/DocumentacionClases_Imexcol.xlsx'
RUTA_SALIDA = 'docs_dominio/DocumentacionClases_Imexcol_v2.xlsx'
SRC_ROOT = 'src/main/java'


def descripcion_por_paquete(nombre, paquete):
    """Genera una descripción breve según paquete y nombre."""
    p = paquete

    if p.endswith('.controlador'):
        if nombre == 'AuthControlador':
            return 'Controlador REST de autenticación de clientes y administradores.'
        base = nombre.replace('Controlador', '')
        return f'Controlador REST para la gestión de la entidad {base}.'
    if p.endswith('.controlador.dto'):
        return f'DTO transversal de la capa de controladores: {nombre}.'
    if p.endswith('.dto'):
        base = nombre.replace('DTO', '')
        return f'DTO de transferencia para la entidad {base}.'
    if p.endswith('.entidad'):
        base = nombre.replace('Entidad', '')
        return f'Entidad de persistencia para la tabla de {base}.'

    if p.endswith('.negocio.dominio'):
        return f'Objeto de dominio que representa a {nombre.replace("Dominio", "")}.'
    if p.endswith('.negocio.fachada'):
        base = nombre.replace('Fachada', '')
        return f'Fachada de negocio para las operaciones de {base}.'
    if p.endswith('.negocio.fachada.impl'):
        base = nombre.replace('FachadaImpl', '')
        return f'Implementación de la fachada de negocio de {base}.'
    if p.endswith('.negocio.casouso'):
        base = nombre.replace('Negocio', '')
        return f'Caso de uso (interfaz) para las reglas de negocio de {base}.'
    if p.endswith('.negocio.casouso.impl'):
        base = nombre.replace('NegocioImpl', '')
        return f'Implementación de los casos de uso de {base}.'
    if '.negocio.casouso.validador' in p:
        return f'Validador de la regla de negocio: {nombre}.'
    if '.negocio.casouso.regla' in p:
        return f'Regla de negocio reutilizable: {nombre}.'
    if p.endswith('.negocio.assembler.dto'):
        return f'Interfaz para ensamblar DTO ↔ Dominio: {nombre}.'
    if p.endswith('.negocio.assembler.dto.impl'):
        base = nombre.replace('DTOAssembler', '')
        return f'Implementación del ensamblador DTO ↔ Dominio para {base}.'
    if p.endswith('.negocio.assembler.entidad'):
        return f'Interfaz para ensamblar Dominio ↔ Entidad: {nombre}.'
    if p.endswith('.negocio.assembler.entidad.impl'):
        base = nombre.replace('EntidadAssembler', '')
        return f'Implementación del ensamblador Dominio ↔ Entidad para {base}.'

    if p.endswith('.datos.dao'):
        base = nombre.replace('DAO', '')
        return f'Interfaz DAO para la persistencia de {base}.'
    if p.endswith('.datos.dao.impl.sqlserver'):
        base = nombre.replace('SqlServerDAO', '')
        return f'Implementación SQL Server del DAO de {base}.'
    if p.endswith('.datos.dao.impl.mapper'):
        base = nombre.replace('Mapper', '')
        return f'Mapper ResultSet → Entidad para {base}.'
    if p.endswith('.datos.dao.impl.sql'):
        base = nombre.replace('Sql', '')
        return f'Sentencias SQL parametrizadas para {base}.'
    if p.endswith('.datos.factoria'):
        return f'Componente de la fábrica de DAOs: {nombre}.'

    if p.endswith('.transversal'):
        return f'Utilidad transversal: {nombre}.'
    if '.transversal.excepcion' in p:
        return f'Componente transversal de manejo de excepciones: {nombre}.'

    # Raíz / Main
    if nombre == 'ImexcolApplication':
        return 'Clase principal con el punto de entrada de Spring Boot.'

    return f'Clase del paquete {paquete}.'


def listar_clases():
    """Devuelve [(nombre, paquete, estereotipo, descripcion), ...]."""
    filas = []
    for r, _, files in os.walk(SRC_ROOT):
        for f in files:
            if not f.endswith('.java'):
                continue
            nombre = f[:-5]
            paquete = r.replace(SRC_ROOT + '/', '').replace('/', '.')
            full = os.path.join(r, f)

            estereotipo = 'class'
            try:
                with open(full, 'r', encoding='utf-8') as fh:
                    contenido = fh.read()
                if re.search(r'\bpublic\s+interface\s+', contenido):
                    estereotipo = 'interface'
                elif re.search(r'\bpublic\s+enum\s+', contenido):
                    estereotipo = 'enum'
                elif re.search(r'\bpublic\s+abstract\s+class\s+', contenido):
                    estereotipo = 'abstract class'
                elif re.search(r'\bpublic\s+final\s+class\s+', contenido):
                    estereotipo = 'final class'
            except Exception:
                pass

            descripcion = descripcion_por_paquete(nombre, paquete)
            filas.append((nombre, paquete, estereotipo, descripcion))

    # Orden: por paquete y luego por nombre
    filas.sort(key=lambda x: (x[1], x[0]))
    return filas


def desunir_si_combinada(ws, fila, col_idx):
    coord = ws.cell(row=fila, column=col_idx).coordinate
    for rango in list(ws.merged_cells.ranges):
        min_col, min_row, max_col, max_row = range_boundaries(str(rango))
        if min_row <= fila <= max_row and min_col <= col_idx <= max_col:
            ws.unmerge_cells(str(rango))
            return True
    return False


def escribir(ws, fila, col_idx, valor):
    desunir_si_combinada(ws, fila, col_idx)
    ws.cell(row=fila, column=col_idx).value = valor


def main():
    print(f'Cargando {RUTA_ENTRADA}…')
    wb = load_workbook(RUTA_ENTRADA)
    print(f'  Hojas: {len(wb.sheetnames)}')

    print('\nLeyendo clases Java…')
    clases = listar_clases()
    print(f'  Total: {len(clases)} clases')

    ws = wb['Clases']
    print(f'\nActualizando hoja "Clases" (estaba con {ws.max_row} filas)…')

    # La cabecera está en R1 — la conservo. Ampliamos a 4 columnas para incluir Paquete.
    escribir(ws, 1, 1, 'Clase')
    escribir(ws, 1, 2, 'Estereotipo')
    escribir(ws, 1, 3, 'Paquete')
    escribir(ws, 1, 4, 'Descripción')

    # Limpiar contenido anterior (R2 en adelante)
    for r in range(2, ws.max_row + 1):
        for c in range(1, max(5, ws.max_column + 1)):
            desunir_si_combinada(ws, r, c)
            ws.cell(row=r, column=c).value = None

    # Escribir las 252 entradas
    for i, (nombre, paquete, estereotipo, descripcion) in enumerate(clases):
        fila = i + 2
        escribir(ws, fila, 1, nombre)
        escribir(ws, fila, 2, estereotipo)
        escribir(ws, fila, 3, paquete)
        escribir(ws, fila, 4, descripcion)

    print(f'  Hoja "Clases" reescrita con {len(clases)} filas (1 cabecera + {len(clases)} datos)')

    print(f'\nGuardando {RUTA_SALIDA}…')
    wb.save(RUTA_SALIDA)
    print('Listo.')


if __name__ == '__main__':
    main()
